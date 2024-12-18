import os
import time
from pathlib import Path
from typing import List
# 测试死区方法

from openai import OpenAI
import openai
import numpy as np
from openpyxl import load_workbook
import shutil
import subprocess
from Fuzz4All.target.target import FResult, Target
import re
import dmPython

# 获取达梦数据库连接
def get_dm_connection():
    # 假设连接配置
    dm_user = 'SYSDBA'
    dm_password = 'SYSDBA'
    dm_host = '127.0.0.1'
    dm_port = '5236'
    return dmPython.connect(user=dm_user, password=dm_password, server=dm_host, port=int(dm_port))

# 获取达梦数据库版本信息
def get_dm_version_info():
    connection = get_dm_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM v$version;")
    version = [row[0] for row in cursor.fetchall()]
    cursor.close()
    connection.close()
    return version

# 获取达梦数据库主要版本号（模拟获取最新 C++版本）
def get_most_recent_dm_version():
    all_versions = get_dm_version_info()
    version_numbers = [re.search(r'(\d+\.\d+\.\d+)', ver).group(1) if re.search(r'(\d+\.\d+\.\d+)', ver) else None for ver in all_versions]
    sorted_versions = sorted([version for version in version_numbers if version is not None], key=lambda x: [int(y) for y in x.split('.')])
    if len(sorted_versions) > 0:
        return sorted_versions[-1]
    else:
        return "no Oracle version found"
MOST_RECENT_DM_VERSION = get_most_recent_dm_version()

class DMTarget(Target):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.folder = "Results/test"
        self.sql = ""
            
        # self.sql_buffer = []
        # self.drop_buffer = []
        self.SYSTEM_MESSAGE = "你是Oracle数据库的安全保卫者，你要不断的发现数据库在触发器的使用过程中可能存在的安全漏洞，并给出对应的测试用例。"
        if kwargs["template"] == "fuzzing_with_config_file":
            config_dict = kwargs["config_dict"]
            self.prompt_used = self._create_prompt_from_config(config_dict)
            self.config_dict = config_dict
        else:
            raise NotImplementedError

    def write_back_file(self, code: str, write_back_name: str = "") -> str:
        # 这里可以将查询写入文件或其他存储位置，为了简化示例，只打印
        """Writes the generated SQL code to a file and returns the file path."""
        if write_back_name != "":
            try:
                with open(write_back_name, "w", encoding="utf-8") as f:
                    f.write(code)
            except:
                pass
        else:
            write_back_name = "tmp/temp{}.fuzz".format(self.CURRENT_TIME)
            try:
                with open(write_back_name, "w", encoding="utf-8") as f:
                    f.write(code)
            except:
                pass
        return write_back_name

    def wrap_prompt(self, prompt: str) -> str:
        """Wraps the prompt in a comment and appends the necessary SQL structure."""
        return f"-- {prompt} \n{self.prompt_used['separator']}\n{self.prompt_used['begin']}"

    def wrap_in_comment(self, prompt: str) -> str:
        """Wraps the given prompt in a SQL comment."""
        return f"-- {prompt}"

    def filter(self, code: str) -> bool:
        """Checks if the generated code contains necessary SQL structures."""
        clean_code = code.replace(self.prompt_used["begin"], "").strip()
        return self.prompt_used["target_api"] in clean_code

    def clean(self, code: str) -> str:
        """Removes comments from the code."""
        # sql, drop_msg = comment_remover(code, lang="dm")
        # return sql
        # code, contrast, drop, trigger = self.extract_sql(code)
        # return "\n".join(code)
        return self.sql
        # return "\n".join(self.sql_buffer)
    
    def clean_code(self, code: str) -> str:
        """Further cleans the code, removing empty lines and unnecessary structures."""
        # code, contrast, drop, trigger = self.extract_sql(code)
        # return "\n".join(code)
        return self.sql
        # return "\n".join(self.sql_buffer)

    def extract_sql(self, sql) -> ([], [], []):
        regex = r"(GRANT|REVOKE|CONNECT|CREATE|INSERT|UPDATE|DROP|SELECT|DELETE|DECLARE|BEGIN|IF|EXECUTE|ELSE|DBMS)\s[^;]*;|(END)[^;]*;"
        pattern = re.compile(regex, re.IGNORECASE | re.MULTILINE)
        matcher = pattern.finditer(sql)

        sql_buffer = [] ## 原始sql
        sql_contrast = [] ## 对比sql
        drop_buffer = [] ## 恢复语句
        cover_list = np.zeros(22) ## 触发器覆盖率
        for match in matcher:
            sql_code = match.group()
            if sql_code.upper().startswith("CREATE"):
                words = sql_code.split()
                if words[1].upper() == "TRIGGER" or (len(words) > 3 and words[3].upper() == "TRIGGER"):
                    cover_list[12] += 1
                    cover_list = self.countTrigger(sql_code, cover_list)
                if (words[1].upper() == "PROCEDURE" or (len(words) > 3 and words[3].upper() == "PROCEDURE")) or (words[1].upper() == "FUNCTION" or (len(words) > 3 and words[3].upper() == "FUNCTION")):
                    if cover_list[13] == 1:
                        cover_list[14] = 1
                    else:
                        cover_list[13] = 1
                    
                if 'ROLE' == words[1].upper():
                    cover_list[21] = 1

                drop_statement = ""
                if "OR" == words[1].upper():
                    drop_statement = ''
                    split1 = words[4].split("(")
                    drop_statement = f"DROP {words[3]} IF EXISTS {split1[0]} CASCADE;"
                    sql_buffer.append(drop_statement)
                else:
                    split1 = words[2].split("(")
                    split = split1[0].split(";")
                    drop_statement = f"DROP {words[1]} IF EXISTS {split[0]} CASCADE;"
                    sql_buffer.append(drop_statement)
                    
                drop_buffer.insert(0, drop_statement)  
            if sql_code.upper().startswith("GRANT"):
                if 'ANY' in sql_code.upper():
                    cover_list[20] = 1
                if 'CONNECT' in sql_code.upper():
                    sql_code = sql_code.replace('CONNECT','CREATE SESSION')
            # if sql_code.upper().startswith("BEGIN") and sql_buffer:
            #     sql_text = sql_buffer.pop()
            #     sql_code = sql_text + sql_code
            if sql_code.upper().startswith("END"):
                if sql_buffer == []:
                    continue
                sql_text = sql_buffer.pop()
                sql_code = sql_text + "\n" + sql_code
            if len(sql_code.upper().split()) == 1:
                continue
            words = sql_code.split()
            sql_buffer.append(sql_code)
        if cover_list[12] > 1:
            if cover_list[13] == 1:
                cover_list[15] = 1
                cover_list[13] = 0
            elif cover_list[14] == 1:
                cover_list[16] = 1
                cover_list[14] = 0
        for list in sql_buffer:
            if "CREATE OR REPLACE TRIGGER" not in list and "CREATE TRIGGER" not in list:
                sql_contrast.append(self.modify_text(list))
        print("sql测例提取结果为：\n","\n".join(sql_buffer))
        print("触发器结果为：{}".format(cover_list))
        return sql_buffer, sql_contrast, drop_buffer, cover_list 
    
    def modify_text(self, text):
        text = text.upper()
        begin_index = text.find('BEGIN')
        if begin_index != -1:
            text = text[:begin_index + 5] + ' \n IF 2 > 1 THEN' + text[begin_index + 5:]
        
        end_index = text.rfind('END')
        if end_index != -1:
            text = text[:end_index] + ' ELSE \n--请在此处插入可以正常执行的代码 \n END IF;\n' + text[end_index:]
        return text

    def countTrigger(self, sql_code, list):
        sql_code = sql_code.upper()
        lines = sql_code.split('\n')
        for line in lines:
            line = line.strip()
            if 'BEFORE' in line:
                list[0] = 1
            elif 'AFTER' in line:
                list[1] = 1
            elif 'INSTAND' in line:
                list[2] = 1
            isMul = False;
            if "OR" in line:
                isMul = True
            if 'BEFORE' in line or 'AFTER' in line or 'INSTAND' in line:
                if "INSERT" in line:
                    if isMul:
                        list[9] = 1
                    else:
                        list[6] = 1
                elif "UPDATE" in line:
                    if isMul:
                        list[10] = 1
                    else:
                        list[7] = 1
                else:
                    if isMul:
                        list[11] = 1
                    else:
                        list[8] = 1
        if 'FOR EACH ROW' in sql_code:
            list[3] = 1
        else:
            list[4] = 1
        if 'EXECUTE IMMEDIATE' in sql_code:
            list[5] = 1
        return list
    
    def countProcedure(self, sql_code, list):
        sql_code = sql_code.upper()
        lines = sql_code.split('\n')
        for line in lines:
            line = line.strip()
            if line.startswith('BEGIN'):
                if 'INSERT' or 'UPDATE' or 'DELETE' or 'SELECT' or 'CREATE' in line:
                    list[17] = 1
                if 'GRANT' or 'REVOKE' in line:
                    list[18] = 1
                if 'EXECUTE IMMEDIATE' in line:
                    list[19] = 1
        return list
    
    def validate_individual(self, filename) -> (FResult, str):
        try:
            with open(filename, "r", encoding="utf-8") as f:
                sql_code = f.read()
            # print(sql_code)
            
            sql_buffer, sql_contrast, drop_buffer, trigger = self.extract_sql(sql_code)
            self.sql = "\n".join(sql_buffer)

            if sql_buffer == [] :
                return FResult.NONE, -1
            else:
                parts = filename.split('/') 
                new_path = 'Results/test/origin/' + parts[len(parts) - 1]
                os.makedirs(os.path.dirname(new_path), exist_ok=True)
                with open(new_path, "w", encoding="utf-8") as f:
                    f.write(sql_code)
                with open(filename, "w", encoding="utf-8") as f:
                    f.write("\n".join(sql_buffer))
                self.update_file(trigger)
            
            success, output = self.execute(sql_buffer)
            if drop_buffer != []:
                self.execute_drop_sql(drop_buffer)
            # if success == False:
            #     return FResult.ERROR, "测试用例执行失败，可能存在语法语义错误。"
            ## 死区生成
            sql_contrast = self.generate_deaf(sql_contrast)
            
            success1, output2 = self.execute(sql_contrast)
            if drop_buffer != []:
                self.execute_drop_sql(drop_buffer)

            # print("触发器执行输出" + str(output) + "\n" + str(output2))
            if success == success1 and success == True:
                if output == output2:
                    return FResult.SAFE, len(output) + 1
                else:
                    return FResult.FAILURE, "触发器前后对比错误，发现漏洞" + "\n对比数据:" + str(output) + "\t" + str(output2)
            else :
                # print("触发器执行失败" + str(output) + "\n" + str(output2))
                return FResult.ERROR, "测试用例执行失败，可能存在语法语义错误。"
        except Exception as e:
            print(e)
            return FResult.ERROR, str(e)
        
    def generate_deaf(self, sql_code) -> []:
        sql = "\n".join(sql_code)
        prompt = "根据上下文，请在我要求的地方插入符合语法规范的任意复杂的SQL语句，不允许使用占位符，不允许添加注释，不允许优化和修改其他地方的SQL，跳过不兼容的语句，不用跟我分析原因直接告诉我结果:"
        sql = prompt + "\n" + sql
        client = OpenAI(
            api_key="sk-67d0f834d27e46eeb819368970a64075", # 如果您没有配置环境变量，请在此处用您的API Key进行替换
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",  # 填写DashScope服务的base_url
        )
        completion = client.chat.completions.create(
            model="qwen2.5-coder-32b-instruct", # 模型列表：https://help.aliyun.com/zh/model-studio/getting-started/models
            messages=[
                {'role': 'system', 'content': '你是Oracle数据库的安全保卫者，你要不断的发现数据库在触发器的使用过程中可能存在的安全漏洞，并给出对应的测试用例。'},
                {'role': 'user', 'content': sql}
            ],
        )
        sql_contrast = self.extract_deaf_sql(completion.choices[0].message.content)
        return sql_contrast

    def extract_deaf_sql(self, sql) -> []:
        regex = r"(GRANT|REVOKE|CONNECT|CREATE|INSERT|UPDATE|DROP|SELECT|DELETE|DECLARE|BEGIN|IF|EXECUTE|ELSE|DBMS)\s[^;]*;|(END)[^;]*;"
        pattern = re.compile(regex, re.IGNORECASE | re.MULTILINE)
        matcher = pattern.finditer(sql)
        sql_buffer = [] ## 原始sql
        for match in matcher:
            sql_code = match.group()
            sql_buffer.append(sql_code)
        return sql_buffer


    def execute(self, sql_code, user="SYSDBA", password="SYSDBA") -> (bool, str):
        connection_string = f"{user}/{password}@{'127.0.0.1'}:{5236}"
        sql_statements = sql_code
        msgs = []
        sql_stmt = []

        for stmt in sql_statements:
            if "CONNECT" in stmt.upper():
                if sql_stmt != []:
                    success, msg = self.execute_sql1("\n".join(sql_stmt), user, password)
                    if not success: return False, msg
                user, password = self._extract_credentials(stmt)
                connection_string = f"{user}/{password}@{'127.0.0.1'}:{5236}"
                sql_stmt = []
                continue
            if stmt.upper().startswith("SELECT"):
                success, msg = self.execute_sql1("\n".join(sql_stmt), user, password)
                if not success: return False, msg
                sql_stmt = []
                success, msg = self.execute_sql(stmt, connection_string)
                if not success:
                    return False, msg
                msgs.append(msg)
            else:
                sql_stmt.append(stmt)
        if sql_stmt != []:
            success, msg = self.execute_sql1("\n".join(sql_stmt), user, password)
            if not success: return False, msg
        return True, "\n".join(msgs)

    def execute_sql(self, sql, connection_string) -> (bool, str):
        try:
            with dmPython.connect(connection_string) as conn:
                with conn.cursor() as cursor:
                    cursor.execute(sql)
                    if cursor.description:  # 查询语句 repr(result_list).lstrip('[').rstrip(']')
                        result_list = [str(row) for row in cursor.fetchall()]
                        return True, str(result_list)
                    else:  # 非查询语句
                        conn.commit()
                        return True, "Success"
        except dmPython.Error as e:
            return False, f"数据库操作出错: {e}"
        
    def execute_drop_sql(self, sql) -> (bool, str):
        try:
            conn = dmPython.connect(user="SYSDBA", password="SYSDBA", server='127.0.0.1', port=int(5236))
            cursor = conn.cursor()
            for stmt in sql:
                cursor.execute(stmt)
                conn.commit()  # 提交非查询语句的事务
            return True, ""
        except dmPython.Error as e:
            return False, f"数据库操作出错: {e}"
        finally:
            try:
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()
            except Exception as close_e:
                return False, f"关闭游标或连接时出错: {close_e}"
        
    def execute_sql1(self, sql_code, user, password) -> (bool, str):
        sql_code += "\n commit;"
        try:
            process = subprocess.Popen(
                ["./disql", user +"/" + password],
                encoding="utf-8",
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                cwd = "/home/useradmin/dmdbms/bin"
            )
        
            stdout, stderr = process.communicate(input=sql_code, timeout=120)
            # print("stdout:", stdout)
            # print("stderr:", stderr)

            if process.returncode == 0 and "错误" not in stdout:
                return True, stdout
            elif process.returncode == 0:
                return False, stdout
            else:
                return False, stderr
        except subprocess.TimeoutExpired:
            return False, "SQL execution timed out"

    def _extract_credentials(self, connect_stmt) -> (str, str):
        _, credentials = connect_stmt.split(maxsplit=1)
        # if credentials
        # if "AS" in credentials:
        #     credentials, _ = credentials.split('AS')
        user, password = credentials.split('/')
        if user == '' or user.upper() == 'SYSTEM' or user.upper() == 'SYS':
            return "SYSDBA", "SYSDBA"
        password = password.split(';')[0]
        # password = password.replace('"', '')
        if password.startswith('"'):
            password = password[1:]
        if password.endswith('"'):
            password = password[:-1]
        return user, password

    def update_file(self, data_array):
        file_path = self.folder + "/count/data.xlsx"
        if not os.path.exists(file_path):
            self.initialize_file()
        workbook = load_workbook(file_path)
        sheet = workbook.active
        next_row_idx = sheet.max_row + 1
        for col_idx, value in enumerate(data_array, start = 1):
            sheet.cell(row = next_row_idx, column = col_idx, value = value)
        workbook.save(file_path)

    def initialize_file(self):
        os.makedirs(self.folder + "/count", exist_ok=True)
        src_file = "data/data.xlsx"
        dst_file = self.folder + "/count/data.xlsx"

        try:
            shutil.copy2(src_file, dst_file)
            print(f"文件 {src_file} 已成功复制到 {dst_file}")
        except FileNotFoundError as e:
            print(f"源文件或目标目录不存在，错误信息：{e}")
        except Exception as e:
            print(f"复制文件时出现其他错误：{e}")