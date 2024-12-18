import pandas as pd
from openpyxl import load_workbook
import random
import numpy as np

# 读取Excel文件，假设文件名为data.xlsx，工作表名为Sheet1（可根据实际情况修改）
df = pd.read_excel("/home/useradmin/LXQ/fuzz4all/data/data.xlsx", sheet_name="Sheet1")

# 通过索引获取列数据，从第三行（索引为2）开始计算到最后一行
column_data = df.iloc[1:, 0]
if column_data.dtype == 'object':
    try:
        # 尝试将数据转换为数值类型
        column_data = pd.to_numeric(column_data)
    except ValueError:
        print("列数据无法转换为数值类型，无法求和")
        raise

column_sum = column_data.sum()
print("该列数据总和为:", column_sum)


# 读取Excel文件，假设文件名为data.xlsx
workbook = load_workbook("/home/useradmin/LXQ/fuzz4all/data/data.xlsx")
# 假设要读取第一个工作表，可根据实际情况修改
sheet = workbook.active
# 获取列数
num_columns = sheet.max_column
print("该Excel文件的列数为:", num_columns)
# new_data = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
new_data = np.zeros(num_columns)
# new_data[0] = 1
# new_data[3] = 1
# new_data[5] = 1
# count1 = random.randint(0, 1)
# if count1 == 1:
#     new_data[7] = 1
# else:
#     new_data[9] = 1
#     new_data[10] = 1
# new_data[21] = 1
# # 获取当前工作表的最后一行的下一行索引
# next_row_idx = sheet.max_row + 1
# for col_idx, value in enumerate(new_data, start = 1):
#     # 给新的一行的每个单元格赋值
#     sheet.cell(row = next_row_idx, column = col_idx, value = value)
# # 保存修改后的Excel文件
# workbook.save("/home/useradmin/LXQ/fuzz4all/data/data.xlsx")
# print("数据已成功添加到最后一行")

def countProcedure(sql_code, list):
    sql_code = sql_code.upper()
    lines = sql_code.split('\n')
    for line in lines:
        line = line.strip()
        if line.startswith('BEGIN'):
            if 'INSERT' or 'UPDATE' or 'DELETE' or 'SELECT' or 'CREATE' in line:
                list[15] = 1
            if 'GRANT' or 'REVOKE' in line:
                list[16] = 1
            if 'EXECUTE IMMEDIATE' in line:
                list[17] = 1
    return list

# sql_code = """create or replace procedure pro_1
# as
# begin
#     insert into table1 values(1,2);
#     grant select on table1 to user1;
# end;
# """

# new_data = countProcedure(sql_code, new_data)

# print(new_data)

count = [[1,2,3,4,5],[2,3,4,5,6]]

sete = [str(row) for row in count]
print(type(sete))
sete = str(sete)
print(type(sete))
print(sete)